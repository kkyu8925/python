from openpyxl import load_workbook

wb = load_workbook("sample.xlsx")
ws = wb.active

# ws.move_range("B1:C11", rows=0, cols=1) # B1~C11 1col 이동
# ws["B1"].value = "국어"  # B1 셀에 '국어' 입력

ws.move_range("C1:C11", rows=5, cols=-1)

wb.save("sample_korean.xlsx")
