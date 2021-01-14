from openpyxl import load_workbook

# wb = load_workbook("sample_fumula.xlsx")
# ws = wb.active

# # 수식 그대로 가져옴
# for row in ws.values:
#     for cell in row:
#         print(cell)

wb = load_workbook("sample_fumula.xlsx", data_only=True)
ws = wb.active

# 수식이 아닌 실제 데이터를 가져옴
# evaluate 되지 않은 상태의 데이터는 None
for row in ws.values:
    for cell in row:
        print(cell)
